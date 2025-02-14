from flask import Flask, request
import json
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime, timedelta
from linebot.v3 import WebhookHandler
from linebot.v3.exceptions import InvalidSignatureError
from linebot.v3.messaging import MessagingApi, Configuration, MessagingApiBlob
from linebot.v3.messaging.api_client import ApiClient
import os
import uuid
from pathlib import Path

app = Flask(__name__)

# LINE Messaging API 的 Token 和 Secret
ACCESS_TOKEN = '1PlQGmb524SP8EccC6ZKIvX47fzf0u9pRZy0E4oCjx71d5gTBTy2U+JzlcfWMc10r4haBWSJHSv7kIE/cnRCnFM6VNtF3CMmTzVAR7n7xtlyiJs3RuuMXhPq+xOv4f9IJontF4iVL8amDiYMJlUxCAdB04t89/1O/w1cDnyilFU='
SECRET = 'ef272fb23654428716e32fefb8f8ba82'

# 初始化 Messaging API 和 WebhookHandler
config = Configuration(access_token=ACCESS_TOKEN)
messaging_api = MessagingApi(api_client=ApiClient(configuration=config))
messaging_api_blob = MessagingApiBlob(api_client=ApiClient(configuration=config))
handler = WebhookHandler(channel_secret=SECRET)

# 文件路径
home = Path.home()
base_path = home / "Downloads" / "行雲流水"

OUTPUT_PATH = base_path / "output_heroku.xlsx"
ID_FILE_PATH = base_path / "id.xlsx"
DOWNLOAD_PATH = base_path / "作業"

# 初始化 Excel 文件（如果不存在）
def initialize_excel():
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws.append(["DateTime (Taiwan)", "User ID", "User Name", "Message Type", "Message Content"])
        for col in ws.columns:
            for cell in col:
                cell.alignment = Alignment(horizontal='center')
        wb.save(OUTPUT_PATH)
    except Exception as e:
        print(f"初始化 Excel 文件失败: {e}")

# 检查文件是否存在，如果不存在则创建
if not os.path.exists(OUTPUT_PATH):
    initialize_excel()

def load_id_mapping(file_path):
    id_mapping = {}
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一列是标题
            user_id = row[2]  # 第 3 列为 id
            part1 = row[3] if row[3] else ""
            part2 = row[4] if row[4] else ""
            custom_name = part1 + "-" + part2
            if user_id and custom_name:  # 确保 id 和显示名称都存在
                id_mapping[user_id] = custom_name
    except Exception as e:
        print(f"加载 id.xlsx 时发生错误: {e}")
    return id_mapping

id_mapping = load_id_mapping(ID_FILE_PATH)

@app.route("/", methods=['POST'])
def linebot():
    body = request.get_data(as_text=True)  # 获取收到的消息内容
    signature = request.headers.get('X-Line-Signature')  # 获取请求头的签名

    try:
        # 验证签名并处理消息
        handler.handle(body=body, signature=signature)

        # 解析事件数据
        json_data = json.loads(body)
        event = json_data['events'][0]
        
        # 获取消息类型
        message_type = event['message']['type']
        message_id = event['message']['id']

        # 获取用户 ID 和显示名称
        user_id = event['source']['userId']
        user_profile = messaging_api.get_profile(user_id)
        user_name = user_profile.display_name
        if user_id in id_mapping:
            file_name = id_mapping[user_id]
        else:
            file_name = user_id

        # 转换时间戳为台湾时间
        timestamp_ms = event['timestamp']  # 时间戳（毫秒）
        timestamp_dt = datetime.utcfromtimestamp(timestamp_ms / 1000) + timedelta(hours=8)
        formatted_time = timestamp_dt.strftime('%Y-%m-%d_%H-%M-%S')

        # 根据消息类型处理
        if message_type == 'text':  # 如果是文本消息
            user_message = event['message']['text']
            message_content = user_message
            message_type_str = 'Text'
        elif message_type == 'sticker':  # 如果是贴图消息
            sticker_id = event['message']['stickerId']
            package_id = event['message']['packageId']
            message_content = f"Sticker ID: {sticker_id}, Package ID: {package_id}"
            message_type_str = 'Sticker'
        elif message_type in ['image', 'file']:  # 如果是图片或檔案消息
            # 獲取文件副檔名
            file_extension = 'jpg' if message_type == 'image' else event['message']['fileName'].split('.')[-1]  # 如果是圖片則副檔名為 jpg，否則就提取原始檔名中的副檔名
            
            unique_id = uuid.uuid4().hex  # 生成唯一識別碼

            # 完整檔案路徑
            download_file_name = f"{file_name}_{formatted_time}_{unique_id}.{file_extension}"
            file_path = os.path.join(DOWNLOAD_PATH, download_file_name)

            # 下载
            try:
                content = messaging_api_blob.get_message_content(message_id)
                with open(file_path, "wb") as f:
                    f.write(content)  # 直接寫入檔案
                '''
                with open(file_path, 'wb') as f:
                    for chunk in content:  # 分塊讀取
                        f.write(chunk)
                '''
                print(f"檔案已儲存至: {file_path}")
            except Exception as e:
                print(f"文件下载失败: {e}")
                
            message_content = f"Downloaded file: {download_file_name}"
            message_type_str = message_type.capitalize()
        else:
            message_content = f"未知消息类型: {message_type}"
            message_type_str = 'Unknown'

        # 打印消息内容、用户名称和 ID
        print(f"时间 (台湾): {formatted_time}")
        print(f"收到的消息类型: {message_type_str}")
        print(f"传讯者名称: {user_name}")
        print(f"传讯者ID: {user_id}")
        print(f"消息内容: {message_content}")

        # 写入 Excel 文件
        try:
            wb = openpyxl.load_workbook(OUTPUT_PATH)
            ws = wb.active
            ws.append([formatted_time, user_id, user_name, message_type_str, message_content])
            wb.save(OUTPUT_PATH)
            print("消息已保存到 Excel 文件。")
        except Exception as e:
            print(f"保存到 Excel 文件失败: {e}")

    except InvalidSignatureError:
        print("Invalid Signature")
    except Exception as e:
        print(f"错误: {e}")
        print(f"收到的内容: {body}")

    return 'OK'

if __name__ == "__main__":
    app.run(debug=True)
